"""Professional Excel export built with openpyxl from Python model outputs."""

from __future__ import annotations

from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .assumptions import Scenario
from .engine import LBOResult


NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
PALE_TAN = "F2E8CF"
LIGHT_GREY = "E7E6E6"
DARK_GREY = "595959"
GREEN = "008000"
RED = "C00000"
WHITE = "FFFFFF"
BLACK = "000000"
INPUT_BLUE = "0000FF"
PASS_GREEN = "E2F0D9"
FAIL_RED = "FCE4D6"

EUR_FORMAT = '€#,##0.0,,"M";(€#,##0.0,,"M");-'
NUMBER_FORMAT = '#,##0.0;(#,##0.0);-'
PERCENT_FORMAT = '0.0%;(0.0%);-'
MULTIPLE_FORMAT = '0.0x;(0.0x);-'
IRR_FORMAT = '0.0%;(0.0%);-'


def _setup_sheet(ws, *, freeze: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 90
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 2
    if freeze:
        ws.freeze_panes = freeze


def _title(ws, title: str, subtitle: str, end_column: int = 12) -> None:
    cell = ws.cell(2, 3, title)
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left")
    ws.cell(3, 3, subtitle).font = Font(
        name="Arial", size=9, italic=True, color=DARK_GREY
    )
    thin = Side(style="thin", color=NAVY)
    for column in range(3, end_column + 1):
        ws.cell(2, column).border = Border(bottom=thin)


def _section(ws, row: int, label: str, start: int, end: int) -> None:
    for column in range(start, end + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.border = Border()
    ws.cell(row, start, label)


def _header_row(ws, row: int, start: int, end: int) -> None:
    for column in range(start, end + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=Side(style="thin", color=WHITE))
    ws.row_dimensions[row].height = 24


def _total_row(ws, row: int, start: int, end: int) -> None:
    for column in range(start, end + 1):
        cell = ws.cell(row, column)
        cell.font = Font(name="Arial", size=10, bold=True, color=BLACK)
        cell.border = Border(top=Side(style="thin", color=BLACK))


def _write_horizontal_schedule(
    ws,
    start_row: int,
    labels_and_values: list[tuple[str, list[float | None], str]],
    years: list[int],
) -> None:
    ws.cell(start_row, 3, "€ in millions except percentages and multiples")
    ws.cell(start_row, 3).font = Font(name="Arial", size=8, italic=True, color=DARK_GREY)
    for offset, year in enumerate(years, 4):
        ws.cell(start_row, offset, f"Year {year}")
    _header_row(ws, start_row, 4, 3 + len(years))
    ws.cell(start_row, 3).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(start_row, 3).font = Font(name="Arial", size=8, bold=True, color=WHITE)
    for row_offset, (label, values, number_format) in enumerate(
        labels_and_values, start_row + 1
    ):
        ws.cell(row_offset, 3, label)
        for column_offset, value in enumerate(values, 4):
            cell = ws.cell(row_offset, column_offset, value)
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right")


def _scenario_summary(results: dict[str, LBOResult]) -> pd.DataFrame:
    rows = []
    for name, result in results.items():
        returns = result.returns
        rows.append(
            {
                "Scenario": name,
                "Revenue Growth": result.scenario.revenue_growth,
                "Exit EBITDA Margin": float(
                    result.operating_model.loc[
                        result.entry.holding_period, "EBITDA Margin"
                    ]
                ),
                "Exit Multiple": float(returns["Exit Multiple"]),
                "Exit Equity Value": float(returns["Sponsor Equity Value"]),
                "Gross Debt Paydown": float(returns["Gross Debt Paydown"]),
                "Exit Net Leverage": float(returns["Exit Net Debt / EBITDA"]),
                "MOIC": float(returns["MOIC"]),
                "IRR": float(returns["IRR"]),
            }
        )
    return pd.DataFrame(rows)


def _build_summary(wb: Workbook, results: dict[str, LBOResult]) -> None:
    ws = wb.create_sheet("Summary")
    _setup_sheet(ws)
    base = results["Base"]
    _title(
        ws,
        "LBO Investment Model",
        f"{base.entry.company} | Fictional transaction | {base.entry.currency} | {base.entry.holding_period}-year hold",
        14,
    )
    summary = _scenario_summary(results)
    _section(ws, 6, "Scenario returns", 3, 11)
    headers = list(summary.columns)
    for column, header in enumerate(headers, 3):
        ws.cell(7, column, header)
    _header_row(ws, 7, 3, 11)
    formats = [None, PERCENT_FORMAT, PERCENT_FORMAT, MULTIPLE_FORMAT, EUR_FORMAT, EUR_FORMAT, MULTIPLE_FORMAT, MULTIPLE_FORMAT, IRR_FORMAT]
    for row, values in enumerate(summary.itertuples(index=False, name=None), 8):
        for column, (value, number_format) in enumerate(zip(values, formats), 3):
            cell = ws.cell(row, column, value)
            if number_format:
                cell.number_format = number_format
        if values[0] == "Base":
            for column in range(3, 12):
                ws.cell(row, column).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                ws.cell(row, column).font = Font(name="Arial", size=10, bold=True)

    _section(ws, 13, "Transaction overview", 3, 6)
    transaction_metrics = [
        ("Entry enterprise value", base.entry.entry_enterprise_value, EUR_FORMAT),
        ("Entry EBITDA", base.entry.entry_ebitda, EUR_FORMAT),
        ("Entry multiple", base.entry.entry_multiple, MULTIPLE_FORMAT),
        ("Initial gross debt", sum(t.leverage_multiple for t in base.debt_assumptions.tranches) * base.entry.entry_ebitda, EUR_FORMAT),
        ("Initial gross leverage", sum(t.leverage_multiple for t in base.debt_assumptions.tranches), MULTIPLE_FORMAT),
        ("Sponsor equity invested", float(base.returns["Entry Equity"]), EUR_FORMAT),
    ]
    for row, (label, value, number_format) in enumerate(transaction_metrics, 14):
        ws.cell(row, 3, label)
        ws.cell(row, 5, value).number_format = number_format

    _section(ws, 13, "Model checks", 8, 11)
    for row, check in enumerate(base.checks.itertuples(index=False), 14):
        ws.cell(row, 8, check.Check)
        status_cell = ws.cell(row, 11, check.Status)
        status_cell.font = Font(name="Arial", size=9, bold=True, color=GREEN if check.Status == "PASS" else RED)
        status_cell.fill = PatternFill("solid", fgColor=PASS_GREEN if check.Status == "PASS" else FAIL_RED)
        status_cell.alignment = Alignment(horizontal="center")

    for column, width in {"C": 24, "D": 15, "E": 17, "F": 17, "G": 17, "H": 25, "I": 17, "J": 17, "K": 18}.items():
        ws.column_dimensions[column].width = width


def _build_assumptions(
    wb: Workbook,
    base: LBOResult,
    scenarios: dict[str, Scenario],
) -> None:
    ws = wb.create_sheet("Assumptions")
    _setup_sheet(ws, freeze="D7")
    ws.sheet_properties.tabColor = "D6B656"
    _title(
        ws,
        "Assumptions",
        "Blue-font cells identify Python inputs; edit assumptions.py and regenerate. All company data are fictional.",
        12,
    )

    _section(ws, 6, "Entry assumptions", 3, 6)
    for column, header in enumerate(["Assumption", "Value", "Units", "Notes"], 3):
        ws.cell(7, column, header)
    _header_row(ws, 7, 3, 6)
    entry_rows = [
        ("Entry revenue", base.entry.entry_revenue, EUR_FORMAT, "LTM revenue"),
        ("Entry EBITDA margin", base.entry.entry_ebitda_margin, PERCENT_FORMAT, "LTM margin"),
        ("Entry EBITDA", base.entry.entry_ebitda, EUR_FORMAT, "Revenue × margin"),
        ("Entry multiple", base.entry.entry_multiple, MULTIPLE_FORMAT, "EV / EBITDA"),
        ("Transaction fees", base.entry.transaction_fee_pct, PERCENT_FORMAT, "% of entry EV"),
        ("Financing fees", base.entry.financing_fee_pct, PERCENT_FORMAT, "% of funded debt"),
        ("Minimum cash", base.entry.minimum_cash, EUR_FORMAT, "Funded at close"),
        ("Exit fees", base.entry.exit_fee_pct, PERCENT_FORMAT, "% of exit EV"),
        ("Holding period", base.entry.holding_period, "0", "Years"),
    ]
    for row, (label, value, number_format, note) in enumerate(entry_rows, 8):
        ws.cell(row, 3, label)
        value_cell = ws.cell(row, 4, value)
        value_cell.number_format = number_format
        value_cell.font = Font(name="Arial", size=10, color=INPUT_BLUE)
        ws.cell(row, 5, "€" if number_format == EUR_FORMAT else "%" if number_format == PERCENT_FORMAT else "x" if number_format == MULTIPLE_FORMAT else "years")
        ws.cell(row, 6, note)

    _section(ws, 19, "Operating assumptions", 3, 6)
    for column, header in enumerate(["Assumption", "Value", "Units", "Notes"], 3):
        ws.cell(20, column, header)
    _header_row(ws, 20, 3, 6)
    operating_rows = [
        ("D&A as % of revenue", base.operating_assumptions.da_pct_revenue, "Non-cash depreciation"),
        ("CapEx as % of revenue", base.operating_assumptions.capex_pct_revenue, "Maintenance and growth CapEx"),
        ("NWC as % of revenue", base.operating_assumptions.nwc_pct_revenue, "Change in annual balance drives cash flow"),
        ("Cash tax rate", base.operating_assumptions.tax_rate, "No NOL carryforward modeled"),
    ]
    for row, (label, value, note) in enumerate(operating_rows, 21):
        ws.cell(row, 3, label)
        cell = ws.cell(row, 4, value)
        cell.number_format = PERCENT_FORMAT
        cell.font = Font(name="Arial", size=10, color=INPUT_BLUE)
        ws.cell(row, 5, "%")
        ws.cell(row, 6, note)

    _section(ws, 28, "Scenario assumptions", 3, 8)
    scenario_headers = ["Scenario", "Revenue Growth", "Annual Margin Expansion", "Exit Multiple", "Cash Sweep", "Capital Structure"]
    for column, header in enumerate(scenario_headers, 3):
        ws.cell(29, column, header)
    _header_row(ws, 29, 3, 8)
    for row, scenario in enumerate(scenarios.values(), 30):
        values = [scenario.name, scenario.revenue_growth, scenario.annual_margin_expansion, scenario.exit_multiple, scenario.cash_sweep_pct, "Held constant"]
        for column, value in enumerate(values, 3):
            cell = ws.cell(row, column, value)
            if column in (4, 5, 7):
                cell.number_format = PERCENT_FORMAT
                cell.font = Font(name="Arial", size=10, color=INPUT_BLUE)
            elif column == 6:
                cell.number_format = MULTIPLE_FORMAT
                cell.font = Font(name="Arial", size=10, color=INPUT_BLUE)

    _section(ws, 36, "Debt tranche assumptions", 3, 10)
    debt_headers = ["Tranche", "Leverage", "Cash Rate", "PIK Rate", "Mandatory Amort.", "Sweep Eligible", "Sweep Priority", "Opening Amount"]
    for column, header in enumerate(debt_headers, 3):
        ws.cell(37, column, header)
    _header_row(ws, 37, 3, 10)
    for row, tranche in enumerate(base.debt_assumptions.tranches, 38):
        values = [tranche.name, tranche.leverage_multiple, tranche.cash_interest_rate, tranche.pik_interest_rate, tranche.mandatory_amortization_pct, "Yes" if tranche.cash_sweep_eligible else "No", tranche.sweep_priority, tranche.leverage_multiple * base.entry.entry_ebitda]
        for column, value in enumerate(values, 3):
            cell = ws.cell(row, column, value)
            if column == 4:
                cell.number_format = MULTIPLE_FORMAT
            elif column in (5, 6, 7):
                cell.number_format = PERCENT_FORMAT
            elif column == 10:
                cell.number_format = EUR_FORMAT
            if column in (4, 5, 6, 7, 8, 9):
                cell.font = Font(name="Arial", size=10, color=INPUT_BLUE)

    widths = {"C": 28, "D": 18, "E": 24, "F": 20, "G": 18, "H": 20, "I": 15, "J": 19}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _build_operating_model(wb: Workbook, base: LBOResult) -> None:
    ws = wb.create_sheet("Operating Model")
    _setup_sheet(ws, freeze="D8")
    _title(
        ws,
        "Operating Model",
        f"Base case | {base.entry.company} | Annual forecast",
        10,
    )
    operating = base.operating_model
    debt = base.debt_schedule
    years = list(operating.index)
    _section(ws, 6, "Income statement and operating cash flow", 3, 3 + len(years))
    labels_and_values = [
        ("Revenue", operating["Revenue"].tolist(), EUR_FORMAT),
        ("Revenue Growth", operating["Revenue Growth"].tolist(), PERCENT_FORMAT),
        ("EBITDA", operating["EBITDA"].tolist(), EUR_FORMAT),
        ("EBITDA Margin", operating["EBITDA Margin"].tolist(), PERCENT_FORMAT),
        ("D&A", operating["D&A"].tolist(), EUR_FORMAT),
        ("EBIT", operating["EBIT"].tolist(), EUR_FORMAT),
        ("Cash Interest", [None] + debt["Cash Interest"].tolist(), EUR_FORMAT),
        ("PIK Interest", [None] + debt["PIK Interest"].tolist(), EUR_FORMAT),
        ("EBT", [None] + debt["EBT"].tolist(), EUR_FORMAT),
        ("Cash Taxes", [None] + debt["Cash Taxes"].tolist(), EUR_FORMAT),
        ("CapEx", operating["CapEx"].tolist(), EUR_FORMAT),
        ("NWC", operating["NWC"].tolist(), EUR_FORMAT),
        ("Change in NWC", operating["Change in NWC"].tolist(), EUR_FORMAT),
        ("FCF Before Debt Paydown", [None] + debt["FCF Before Debt Paydown"].tolist(), EUR_FORMAT),
    ]
    _write_horizontal_schedule(ws, 7, labels_and_values, years)
    for row in (10, 13, 17, 21):
        _total_row(ws, row, 3, 3 + len(years))
    ws.column_dimensions["C"].width = 28
    for column in range(4, 4 + len(years)):
        ws.column_dimensions[get_column_letter(column)].width = 14

    chart = LineChart()
    chart.title = "Revenue and EBITDA evolution"
    chart.style = 13
    chart.y_axis.title = "€ millions"
    chart.x_axis.title = "Year"
    chart.height = 8.0
    chart.width = 15.5
    categories = Reference(ws, min_col=4, max_col=3 + len(years), min_row=7)
    for source_row in (8, 10):
        data = Reference(
            ws,
            min_col=3,
            max_col=3 + len(years),
            min_row=source_row,
            max_row=source_row,
        )
        chart.add_data(data, titles_from_data=True, from_rows=True)
    chart.set_categories(categories)
    chart.legend.position = "t"
    ws.add_chart(chart, "C25")

    margin_chart = LineChart()
    margin_chart.title = "EBITDA margin evolution"
    margin_chart.style = 13
    margin_chart.y_axis.title = "Margin"
    margin_chart.y_axis.numFmt = "0.0%"
    margin_chart.height = 8.0
    margin_chart.width = 15.5
    margin_data = Reference(ws, min_col=3, max_col=3 + len(years), min_row=11, max_row=11)
    margin_chart.add_data(margin_data, titles_from_data=True, from_rows=True)
    margin_chart.set_categories(categories)
    margin_chart.legend = None
    ws.add_chart(margin_chart, "K25")


def _build_debt_schedule(wb: Workbook, base: LBOResult) -> None:
    ws = wb.create_sheet("Debt Schedule")
    _setup_sheet(ws, freeze="D8")
    _title(
        ws,
        "Debt Schedule",
        "Cash interest on average pre-sweep debt; PIK accrues on opening principal; sweep applied at year-end.",
        10,
    )
    debt = base.debt_schedule
    years = list(debt.index)
    current_row = 6
    for tranche in base.debt_assumptions.tranches:
        _section(ws, current_row, tranche.name, 3, 3 + len(years))
        labels_and_values = [
            ("Opening Debt", debt[f"{tranche.name} Opening"].tolist(), EUR_FORMAT),
            ("Cash Interest", debt[f"{tranche.name} Cash Interest"].tolist(), EUR_FORMAT),
            ("PIK Interest", debt[f"{tranche.name} PIK Interest"].tolist(), EUR_FORMAT),
            ("Mandatory Amortization", debt[f"{tranche.name} Mandatory Amortization"].tolist(), EUR_FORMAT),
            ("Cash Sweep", debt[f"{tranche.name} Cash Sweep"].tolist(), EUR_FORMAT),
            ("Closing Debt", debt[f"{tranche.name} Closing"].tolist(), EUR_FORMAT),
        ]
        _write_horizontal_schedule(ws, current_row + 1, labels_and_values, years)
        _total_row(ws, current_row + 7, 3, 3 + len(years))
        current_row += 10

    _section(ws, current_row, "Cash flow and leverage", 3, 3 + len(years))
    combined_rows = [
        ("Opening Cash", debt["Opening Cash"].tolist(), EUR_FORMAT),
        ("FCF Before Debt Paydown", debt["FCF Before Debt Paydown"].tolist(), EUR_FORMAT),
        ("Mandatory Amortization", debt["Mandatory Amortization"].tolist(), EUR_FORMAT),
        ("Cash Before Sweep", debt["Cash Before Sweep"].tolist(), EUR_FORMAT),
        ("Cash Sweep", debt["Cash Sweep"].tolist(), EUR_FORMAT),
        ("Closing Cash", debt["Closing Cash"].tolist(), EUR_FORMAT),
        ("Total Debt", debt["Total Debt"].tolist(), EUR_FORMAT),
        ("Net Debt", debt["Net Debt"].tolist(), EUR_FORMAT),
        ("Net Debt / EBITDA", debt["Net Debt / EBITDA"].tolist(), MULTIPLE_FORMAT),
    ]
    _write_horizontal_schedule(ws, current_row + 1, combined_rows, years)
    for row in (current_row + 7, current_row + 8, current_row + 9, current_row + 10):
        _total_row(ws, row, 3, 3 + len(years))
    ws.column_dimensions["C"].width = 29
    for column in range(4, 4 + len(years)):
        ws.column_dimensions[get_column_letter(column)].width = 15

    chart_row = current_row + 13
    debt_chart = BarChart()
    debt_chart.type = "col"
    debt_chart.grouping = "stacked"
    debt_chart.overlap = 100
    debt_chart.title = "Debt paydown by tranche"
    debt_chart.y_axis.title = "€ millions"
    debt_chart.height = 8.0
    debt_chart.width = 15.5
    tranche_closing_rows = [13, 23, 33]
    for tranche, closing_row in zip(
        base.debt_assumptions.tranches, tranche_closing_rows
    ):
        values = Reference(
            ws,
            min_col=4,
            max_col=3 + len(years),
            min_row=closing_row,
            max_row=closing_row,
        )
        debt_chart.append(Series(values, title=tranche.name))
    categories = Reference(ws, min_col=4, max_col=3 + len(years), min_row=7)
    debt_chart.set_categories(categories)
    debt_chart.legend.position = "t"
    ws.add_chart(debt_chart, f"C{chart_row}")

    leverage_row = current_row + 10
    leverage_chart = LineChart()
    leverage_chart.title = "Net Debt / EBITDA"
    leverage_chart.y_axis.title = "x"
    leverage_chart.height = 8.0
    leverage_chart.width = 15.5
    values = Reference(ws, min_col=3, max_col=3 + len(years), min_row=leverage_row, max_row=leverage_row)
    leverage_chart.add_data(values, titles_from_data=True, from_rows=True)
    leverage_chart.set_categories(categories)
    leverage_chart.legend = None
    ws.add_chart(leverage_chart, f"K{chart_row}")


def _build_returns(wb: Workbook, results: dict[str, LBOResult]) -> None:
    ws = wb.create_sheet("Returns")
    _setup_sheet(ws)
    base = results["Base"]
    _title(ws, "Returns", "Sources & Uses, exit bridge and sponsor returns", 12)

    _section(ws, 6, "Sources & Uses", 3, 8)
    uses = base.sources_uses.loc[base.sources_uses["Type"] == "Use"]
    sources = base.sources_uses.loc[base.sources_uses["Type"] == "Source"]
    ws.cell(7, 3, "Uses")
    ws.cell(7, 4, "Amount")
    ws.cell(7, 6, "Sources")
    ws.cell(7, 7, "Amount")
    _header_row(ws, 7, 3, 4)
    _header_row(ws, 7, 6, 7)
    max_rows = max(len(uses), len(sources))
    for offset in range(max_rows):
        row = 8 + offset
        if offset < len(uses):
            ws.cell(row, 3, uses.iloc[offset]["Item"])
            ws.cell(row, 4, float(uses.iloc[offset]["Amount"])).number_format = EUR_FORMAT
        if offset < len(sources):
            ws.cell(row, 6, sources.iloc[offset]["Item"])
            ws.cell(row, 7, float(sources.iloc[offset]["Amount"])).number_format = EUR_FORMAT
    total_row = 8 + max_rows
    ws.cell(total_row, 3, "Total Uses")
    ws.cell(total_row, 4, float(uses["Amount"].sum())).number_format = EUR_FORMAT
    ws.cell(total_row, 6, "Total Sources")
    ws.cell(total_row, 7, float(sources["Amount"].sum())).number_format = EUR_FORMAT
    _total_row(ws, total_row, 3, 4)
    _total_row(ws, total_row, 6, 7)

    _section(ws, 16, "Scenario returns", 3, 11)
    summary = _scenario_summary(results)
    headers = ["Scenario", "Exit EBITDA", "Exit Multiple", "Exit EV", "Exit Debt", "Exit Cash", "Exit Equity", "MOIC", "IRR"]
    for column, header in enumerate(headers, 3):
        ws.cell(17, column, header)
    _header_row(ws, 17, 3, 11)
    for row, (name, result) in enumerate(results.items(), 18):
        returns = result.returns
        values = [name, float(returns["Exit EBITDA"]), float(returns["Exit Multiple"]), float(returns["Exit Enterprise Value"]), float(returns["Less: Closing Debt"]), float(returns["Add: Closing Cash"]), float(returns["Sponsor Equity Value"]), float(returns["MOIC"]), float(returns["IRR"])]
        for column, value in enumerate(values, 3):
            cell = ws.cell(row, column, value)
            if column in (4, 6, 7, 8, 9):
                cell.number_format = EUR_FORMAT
            elif column in (5, 10):
                cell.number_format = MULTIPLE_FORMAT
            elif column == 11:
                cell.number_format = IRR_FORMAT
        if name == "Base":
            for column in range(3, 12):
                ws.cell(row, column).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    _section(ws, 23, "Base case exit equity bridge", 3, 7)
    exit_rows = [
        ("Exit EBITDA", base.returns["Exit EBITDA"], EUR_FORMAT),
        ("Exit Multiple", base.returns["Exit Multiple"], MULTIPLE_FORMAT),
        ("Exit Enterprise Value", base.returns["Exit Enterprise Value"], EUR_FORMAT),
        ("Less: Closing Debt", -float(base.returns["Less: Closing Debt"]), EUR_FORMAT),
        ("Add: Closing Cash", base.returns["Add: Closing Cash"], EUR_FORMAT),
        ("Less: Exit Fees", -float(base.returns["Less: Exit Fees"]), EUR_FORMAT),
        ("Sponsor Equity Value", base.returns["Sponsor Equity Value"], EUR_FORMAT),
        ("MOIC", base.returns["MOIC"], MULTIPLE_FORMAT),
        ("IRR", base.returns["IRR"], IRR_FORMAT),
    ]
    for row, (label, value, number_format) in enumerate(exit_rows, 24):
        ws.cell(row, 3, label)
        ws.cell(row, 5, float(value)).number_format = number_format
    _total_row(ws, 30, 3, 5)

    _section(ws, 23, "Base case value creation", 8, 11)
    for row, item in enumerate(base.value_creation.itertuples(index=False), 24):
        ws.cell(row, 8, item.Component)
        ws.cell(row, 11, item.Amount).number_format = EUR_FORMAT
    _total_row(ws, 31, 8, 11)

    equity_chart = BarChart()
    equity_chart.type = "col"
    equity_chart.title = "Sponsor equity value"
    equity_chart.y_axis.title = "€ millions"
    equity_chart.height = 8.0
    equity_chart.width = 15.5
    _section(ws, 35, "Sponsor equity value by scenario", 3, 5)
    chart_data_start = 36
    ws.cell(chart_data_start, 3, "Scenario")
    ws.cell(chart_data_start, 4, "Entry Equity")
    ws.cell(chart_data_start, 5, "Exit Equity")
    for row, (name, result) in enumerate(results.items(), chart_data_start + 1):
        ws.cell(row, 3, name)
        ws.cell(row, 4, float(result.returns["Entry Equity"]))
        ws.cell(row, 5, float(result.returns["Sponsor Equity Value"]))
    data = Reference(ws, min_col=4, max_col=5, min_row=chart_data_start, max_row=chart_data_start + len(results))
    categories = Reference(ws, min_col=3, min_row=chart_data_start + 1, max_row=chart_data_start + len(results))
    equity_chart.add_data(data, titles_from_data=True)
    equity_chart.set_categories(categories)
    equity_chart.legend.position = "t"
    ws.add_chart(equity_chart, "C42")
    _header_row(ws, chart_data_start, 3, 5)
    for row in range(chart_data_start + 1, chart_data_start + len(results) + 1):
        ws.cell(row, 4).number_format = EUR_FORMAT
        ws.cell(row, 5).number_format = EUR_FORMAT

    for column, width in {"C": 31, "D": 17, "E": 17, "F": 29, "G": 17, "H": 31, "I": 17, "J": 17, "K": 19}.items():
        ws.column_dimensions[column].width = width


def _write_sensitivity_grid(ws, start_row: int, start_col: int, title: str, frame: pd.DataFrame, index_format: str) -> int:
    end_col = start_col + len(frame.columns)
    _section(ws, start_row, title, start_col, end_col)
    ws.cell(start_row + 1, start_col, frame.index.name or "Driver")
    for offset, value in enumerate(frame.columns, start_col + 1):
        ws.cell(start_row + 1, offset, value).number_format = MULTIPLE_FORMAT
    _header_row(ws, start_row + 1, start_col, end_col)
    for row_offset, (index_value, row_values) in enumerate(frame.iterrows(), start_row + 2):
        ws.cell(row_offset, start_col, index_value).number_format = index_format
        for column_offset, value in enumerate(row_values, start_col + 1):
            ws.cell(row_offset, column_offset, float(value)).number_format = IRR_FORMAT
    data_range = f"{get_column_letter(start_col + 1)}{start_row + 2}:{get_column_letter(end_col)}{start_row + 1 + len(frame)}"
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    return start_row + len(frame) + 3


def _build_sensitivities(wb: Workbook, sensitivities: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Sensitivities")
    _setup_sheet(ws, freeze="D8")
    _title(ws, "Sensitivities", "Each cell re-runs operating cash flow, debt paydown and exit returns.", 12)
    next_row = _write_sensitivity_grid(ws, 6, 3, "IRR: Entry Multiple × Exit Multiple", sensitivities["Entry x Exit IRR"], MULTIPLE_FORMAT)
    next_row = _write_sensitivity_grid(ws, next_row + 2, 3, "IRR: Exit Multiple × Exit EBITDA CAGR", sensitivities["Exit x EBITDA Growth IRR"], PERCENT_FORMAT)

    holding = sensitivities["Holding Period"]
    start_row = next_row + 2
    _section(ws, start_row, "Holding period sensitivity", 3, 8)
    headers = ["Holding Period", "Exit EBITDA", "Exit Equity Value", "MOIC", "IRR", "Exit Net Leverage"]
    for column, header in enumerate(headers, 3):
        ws.cell(start_row + 1, column, header)
    _header_row(ws, start_row + 1, 3, 8)
    for row, (period, values) in enumerate(holding.iterrows(), start_row + 2):
        row_values = [period, values["Exit EBITDA"], values["Exit Equity Value"], values["MOIC"], values["IRR"], values["Exit Net Debt / EBITDA"]]
        for column, value in enumerate(row_values, 3):
            cell = ws.cell(row, column, float(value))
            if column in (4, 5):
                cell.number_format = EUR_FORMAT
            elif column in (6, 8):
                cell.number_format = MULTIPLE_FORMAT
            elif column == 7:
                cell.number_format = IRR_FORMAT
            else:
                cell.number_format = "0"
    ws.column_dimensions["C"].width = 24
    for column in range(4, 9):
        ws.column_dimensions[get_column_letter(column)].width = 19


def _finalize_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    font = copy(cell.font)
                    font.name = "Arial"
                    if font.sz is None:
                        font.sz = 10
                    cell.font = font
        ws.sheet_view.selection[0].activeCell = "C2"
        ws.sheet_view.selection[0].sqref = "C2"
        ws.auto_filter.ref = None
        ws.oddFooter.center.text = "Fictional transaction — analytical and portfolio use only"
        ws.oddFooter.center.size = 8
        ws.oddFooter.center.color = DARK_GREY


def export_workbook(
    results: dict[str, LBOResult],
    scenarios: dict[str, Scenario],
    sensitivities: dict[str, pd.DataFrame],
    output_path: str | Path,
) -> Path:
    """Create the six-sheet Excel model from calculated Python schedules."""

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    _build_summary(wb, results)
    _build_assumptions(wb, results["Base"], scenarios)
    _build_operating_model(wb, results["Base"])
    _build_debt_schedule(wb, results["Base"])
    _build_returns(wb, results)
    _build_sensitivities(wb, sensitivities)
    _finalize_workbook(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)
    return path


def validate_workbook(output_path: str | Path) -> dict[str, object]:
    """Re-open the exported workbook and run structural quality checks."""

    workbook = load_workbook(output_path, data_only=False)
    expected_sheets = [
        "Summary",
        "Assumptions",
        "Operating Model",
        "Debt Schedule",
        "Returns",
        "Sensitivities",
    ]
    formula_errors = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(
                    error in cell.value
                    for error in (
                        "#REF!",
                        "#DIV/0!",
                        "#VALUE!",
                        "#NAME?",
                        "#N/A",
                        "#NUM!",
                    )
                ):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}: {cell.value}")
    result = {
        "sheet_order": workbook.sheetnames,
        "sheet_order_matches": workbook.sheetnames == expected_sheets,
        "formula_errors": formula_errors,
        "chart_counts": {ws.title: len(ws._charts) for ws in workbook.worksheets},
    }
    workbook.close()
    return result
