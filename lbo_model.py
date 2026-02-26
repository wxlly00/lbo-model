"""
LBO Model — TechRetail SA
Author: Wilfried LAWSON HELLU | Finance Analyst
GitHub: github.com/Wxlly00

End-to-end Leveraged Buyout model:
- Income statement projections (5 years)
- Debt schedule (Senior + Mezzanine)
- Cash flow model
- Multi-scenario exit analysis (IRR, MOIC, DPI)
"""

import numpy as np
import pandas as pd
from tabulate import tabulate
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# MODEL INPUTS
# ─────────────────────────────────────────────
COMPANY = "TechRetail SA"
ENTRY_EV = 450_000_000          # EUR 450M
ENTRY_EBITDA = 60_000_000       # EUR 60M (7.5x multiple)
ENTRY_MULTIPLE = 7.5
REVENUE_Y0 = 300_000_000        # EUR 300M
EBITDA_MARGIN_Y0 = 0.20         # 20%
REVENUE_CAGR = 0.08             # 8% annual growth
MARGIN_EXPANSION = 0.01         # +1% per year
DA_PCT = 0.04                   # D&A as % of revenue
CAPEX_PCT = 0.03                # Capex as % of revenue
WC_CHANGE_PCT = 0.01            # NWC increase as % of revenue
TAX_RATE = 0.28                 # 28% corporate tax
HOLDING_YEARS = 5

# Debt structure
SENIOR_DEBT = 247_500_000       # 55% of EV @ 6%
SENIOR_RATE = 0.06
SENIOR_AMORT_PCT = 0.05         # 5% mandatory annual amortization
MEZZ_DEBT = 67_500_000          # 15% of EV @ 10% PIK
MEZZ_RATE = 0.10
EQUITY = 135_000_000            # 30% of EV

# Exit scenarios
EXIT_SCENARIOS = {
    "Bear Case": 6.0,
    "Base Case": 7.5,
    "Bull Case": 9.0,
}


def fmt(n: float, currency: bool = True) -> str:
    """Format large numbers in EUR millions."""
    if currency:
        return f"€{n/1e6:.1f}M"
    return f"{n:.1f}%"


def build_income_statement() -> pd.DataFrame:
    """Build 5-year projected income statement."""
    rows = []
    for year in range(1, HOLDING_YEARS + 1):
        revenue = REVENUE_Y0 * (1 + REVENUE_CAGR) ** year
        ebitda_margin = EBITDA_MARGIN_Y0 + MARGIN_EXPANSION * year
        ebitda = revenue * ebitda_margin
        da = revenue * DA_PCT
        ebit = ebitda - da
        rows.append({
            "Year": year,
            "Revenue (€M)": revenue / 1e6,
            "Revenue Growth": f"{REVENUE_CAGR*100:.0f}%",
            "EBITDA (€M)": ebitda / 1e6,
            "EBITDA Margin": f"{ebitda_margin*100:.0f}%",
            "D&A (€M)": da / 1e6,
            "EBIT (€M)": ebit / 1e6,
        })
    return pd.DataFrame(rows).set_index("Year")


def build_debt_schedule(income_stmt: pd.DataFrame) -> pd.DataFrame:
    """Build year-by-year debt amortization schedule."""
    rows = []
    senior_bal = SENIOR_DEBT
    mezz_bal = MEZZ_DEBT

    for year in range(1, HOLDING_YEARS + 1):
        # Senior mandatory amortization
        senior_amort = min(SENIOR_DEBT * SENIOR_AMORT_PCT, senior_bal)
        senior_interest = senior_bal * SENIOR_RATE
        senior_bal_close = senior_bal - senior_amort

        # Mezzanine: PIK (interest accrues to principal)
        mezz_interest_cash = 0
        mezz_accrued = mezz_bal * MEZZ_RATE
        mezz_bal_close = mezz_bal + mezz_accrued

        rows.append({
            "Year": year,
            "Senior Opening (€M)": senior_bal / 1e6,
            "Senior Interest (€M)": senior_interest / 1e6,
            "Senior Amortization (€M)": senior_amort / 1e6,
            "Senior Closing (€M)": senior_bal_close / 1e6,
            "Mezz Opening (€M)": mezz_bal / 1e6,
            "Mezz PIK Accrual (€M)": mezz_accrued / 1e6,
            "Mezz Closing (€M)": mezz_bal_close / 1e6,
        })
        senior_bal = senior_bal_close
        mezz_bal = mezz_bal_close

    return pd.DataFrame(rows).set_index("Year"), senior_bal, mezz_bal


def build_cash_flow(income_stmt: pd.DataFrame, debt_sched: pd.DataFrame) -> pd.DataFrame:
    """Build free cash flow statement."""
    rows = []
    for year in range(1, HOLDING_YEARS + 1):
        revenue = REVENUE_Y0 * (1 + REVENUE_CAGR) ** year
        ebitda = income_stmt.loc[year, "EBITDA (€M)"] * 1e6
        senior_interest = debt_sched.loc[year, "Senior Interest (€M)"] * 1e6
        capex = revenue * CAPEX_PCT
        wc_change = revenue * WC_CHANGE_PCT
        ebt = income_stmt.loc[year, "EBIT (€M)"] * 1e6 - senior_interest
        tax = max(ebt * TAX_RATE, 0)
        net_income = ebt - tax
        fcf = ebitda - capex - wc_change - senior_interest - tax
        rows.append({
            "Year": year,
            "EBITDA (€M)": ebitda / 1e6,
            "- Capex (€M)": -capex / 1e6,
            "- NWC Change (€M)": -wc_change / 1e6,
            "- Senior Interest (€M)": -senior_interest / 1e6,
            "- Taxes (€M)": -tax / 1e6,
            "Free Cash Flow (€M)": fcf / 1e6,
        })
    return pd.DataFrame(rows).set_index("Year")


def calculate_returns(year5_ebitda: float, senior_bal_exit: float, mezz_bal_exit: float) -> dict:
    """Calculate exit returns for each scenario."""
    results = {}
    for scenario, exit_mult in EXIT_SCENARIOS.items():
        exit_ev = year5_ebitda * exit_mult
        total_debt = senior_bal_exit + mezz_bal_exit
        equity_value = exit_ev - total_debt
        moic = equity_value / EQUITY
        irr = (equity_value / EQUITY) ** (1 / HOLDING_YEARS) - 1
        results[scenario] = {
            "Exit Multiple": f"{exit_mult:.1f}x",
            "Exit EV (€M)": f"€{exit_ev/1e6:.0f}M",
            "Total Debt at Exit (€M)": f"€{total_debt/1e6:.0f}M",
            "Equity Value (€M)": f"€{equity_value/1e6:.0f}M",
            "Entry Equity (€M)": f"€{EQUITY/1e6:.0f}M",
            "MOIC": f"{moic:.2f}x",
            "IRR": f"{irr*100:.1f}%",
            "DPI": f"{moic:.2f}x",
        }
    return results


def save_to_excel(income_stmt, debt_sched, returns_df):
    """Save model outputs to formatted Excel file."""
    wb = openpyxl.Workbook()
    
    # Colors
    header_fill = PatternFill("solid", fgColor="0A1628")
    gold_fill = PatternFill("solid", fgColor="C9A84C")
    alt_fill = PatternFill("solid", fgColor="EFF2F7")
    
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    title_font = Font(color="0A1628", bold=True, name="Calibri", size=12)
    
    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1: Income Statement
    ws1 = wb.active
    ws1.title = "Income Statement"
    ws1.cell(1, 1, f"LBO MODEL — {COMPANY} | Income Statement Projections").font = title_font
    ws1.cell(2, 1, f"Entry EV: €{ENTRY_EV/1e6:.0f}M | Entry Multiple: {ENTRY_MULTIPLE}x | Holding: {HOLDING_YEARS} years")
    
    headers = ["Year"] + list(income_stmt.columns)
    for j, h in enumerate(headers, 1):
        ws1.cell(4, j, h)
    style_header_row(ws1, 4, len(headers))
    
    for i, (year, row) in enumerate(income_stmt.iterrows(), 5):
        ws1.cell(i, 1, year)
        for j, val in enumerate(row, 2):
            ws1.cell(i, j, round(val, 2) if isinstance(val, float) else val)
    
    for col in ws1.columns:
        ws1.column_dimensions[get_column_letter(col[0].column)].width = 20

    # Sheet 2: Debt Schedule
    ws2 = wb.create_sheet("Debt Schedule")
    ws2.cell(1, 1, f"Debt Schedule | Senior @ {SENIOR_RATE*100:.0f}% + Mezzanine PIK @ {MEZZ_RATE*100:.0f}%").font = title_font
    
    headers2 = ["Year"] + list(debt_sched.columns)
    for j, h in enumerate(headers2, 1):
        ws2.cell(3, j, h)
    style_header_row(ws2, 3, len(headers2))
    
    for i, (year, row) in enumerate(debt_sched.iterrows(), 4):
        ws2.cell(i, 1, year)
        for j, val in enumerate(row, 2):
            ws2.cell(i, j, round(val, 2) if isinstance(val, float) else val)
    
    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 22

    # Sheet 3: Returns Analysis
    ws3 = wb.create_sheet("Returns Analysis")
    ws3.cell(1, 1, "Exit Returns Analysis | Bear / Base / Bull Scenarios").font = title_font
    
    headers3 = ["Metric"] + list(returns_df.keys())
    for j, h in enumerate(headers3, 1):
        ws3.cell(3, j, h)
    style_header_row(ws3, 3, len(headers3))
    
    metrics = list(list(returns_df.values())[0].keys())
    for i, metric in enumerate(metrics, 4):
        ws3.cell(i, 1, metric)
        for j, scenario in enumerate(returns_df.values(), 2):
            ws3.cell(i, j, scenario[metric])
    
    for col in ws3.columns:
        ws3.column_dimensions[get_column_letter(col[0].column)].width = 22
    
    wb.save("lbo_output.xlsx")
    print("\n✓ Excel file saved: lbo_output.xlsx")


def main():
    print("=" * 70)
    print(f"  LBO MODEL — {COMPANY}")
    print(f"  Entry EV: €{ENTRY_EV/1e6:.0f}M | Multiple: {ENTRY_MULTIPLE}x EBITDA")
    print(f"  Debt: Senior €{SENIOR_DEBT/1e6:.0f}M ({SENIOR_RATE*100:.0f}%) + Mezz €{MEZZ_DEBT/1e6:.0f}M (PIK {MEZZ_RATE*100:.0f}%)")
    print(f"  Equity: €{EQUITY/1e6:.0f}M | Holding: {HOLDING_YEARS} years")
    print("=" * 70)

    # Build model
    income_stmt = build_income_statement()
    debt_sched, senior_exit, mezz_exit = build_debt_schedule(income_stmt)
    cash_flow = build_cash_flow(income_stmt, debt_sched)

    print("\n📊 INCOME STATEMENT PROJECTIONS (€M)")
    print(tabulate(income_stmt.round(1), headers="keys", tablefmt="rounded_outline", floatfmt=".1f"))

    print("\n🏦 DEBT SCHEDULE (€M)")
    senior_cols = ["Senior Opening (€M)", "Senior Interest (€M)", "Senior Amortization (€M)", "Senior Closing (€M)"]
    mezz_cols = ["Mezz Opening (€M)", "Mezz PIK Accrual (€M)", "Mezz Closing (€M)"]
    print("  SENIOR TRANCHE:")
    print(tabulate(debt_sched[senior_cols].round(1), headers="keys", tablefmt="rounded_outline", floatfmt=".1f"))
    print("  MEZZANINE TRANCHE (PIK):")
    print(tabulate(debt_sched[mezz_cols].round(1), headers="keys", tablefmt="rounded_outline", floatfmt=".1f"))

    print("\n💵 FREE CASH FLOW (€M)")
    print(tabulate(cash_flow.round(1), headers="keys", tablefmt="rounded_outline", floatfmt=".1f"))

    # Calculate exit returns
    year5_ebitda = income_stmt.loc[5, "EBITDA (€M)"] * 1e6
    returns = calculate_returns(year5_ebitda, senior_exit, mezz_exit)

    print("\n🎯 EXIT RETURNS ANALYSIS")
    print(f"  Year 5 EBITDA: €{year5_ebitda/1e6:.1f}M")
    print(f"  Senior Debt at Exit: €{senior_exit/1e6:.1f}M | Mezz (PIK): €{mezz_exit/1e6:.1f}M")
    
    returns_table = pd.DataFrame(returns).T
    print(tabulate(returns_table, headers="keys", tablefmt="rounded_outline"))

    # Save Excel
    save_to_excel(income_stmt, debt_sched, returns)

    print("\n✅ LBO Model completed successfully!")
    print("   Author: Wilfried LAWSON HELLU | github.com/Wxlly00")


if __name__ == "__main__":
    main()
